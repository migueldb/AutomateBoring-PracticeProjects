#! python3
# Automate Boring Stuff Practice Project - Chapter 12
# cellInverter.py - transposes the entire spreadsheet contents
#
# NOTE: This script requries a workbook saved in the working directory
#

import openpyxl as o

def invertCells(_sheetNumber_,_fileName_):
  # Create workbook object
  wb = o.load_workbook(_fileName_)
  # Select the given worksheet by number
  sheet = wb.worksheets[_sheetNumber_-1]
  # Create a new workbook object to store the transposed spreadsheet
  newWb = o.Workbook()
  newSheet = newWb.active
  contents = []

  # Loop through columns
  for x in range (sheet.max_column):
    contents.append([])       # Finally, this is how you add lists inside a list
    # Loop through rows
    for y in range (sheet.max_row):
      # Read cell by cell and save the contents switching rows with columns
      contents[x].append(sheet.cell(row=y+1,column=x+1).value)
      newSheet.cell(row=x+1,column=y+1,value=contents[x][y])
  # Save new file with a different name
  newWb.save('inv'+_fileName_)
  

sheetNumber = 1
# User can insert the file name here
fileName = 'insertmultiplicationTable6.xlsx'
invertCells(sheetNumber,fileName)
