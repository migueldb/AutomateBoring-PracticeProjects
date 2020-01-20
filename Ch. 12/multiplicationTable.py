#! python3
# Automate Boring Stuff Practice Project - Chapter 12
# multiplicationTable.py - this program creates a multiplication table from a 
#                          given number, n.  The size of the table is n x n.
#
# NOTE: This script requires numpy module. Please check
# https://scipy.org/install.html
# for instructions
#
# USAGE: From multiplicationTable.py directory type on the command prompt:
# C:\path\to\python\py.exe multiplicationTable.py <table-size>, i.e.
# py multiplicationTable.py 6


import openpyxl as o
import numpy    as n
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter, \
column_index_from_string # openpyxl.cell has been superseded
import sys


def createMultTable(_number_):
  # Create workbook object
  wb = o.Workbook()
  # Select active worksheet
  sheet = wb.active
  # Create arrays with integers from 1 to size
  a = n.arange(1,_number_+1).reshape(_number_,1)
  b = a.reshape(1,_number_)
  # Numpy dot product of two arrays
  c = a@b
  # print(c)

  it = n.nditer(c, flags=['multi_index']) # this is how to iterate with arrays
  # Formatting
  fontObj1 = Font(bold=True)
  # Create first row and first column with integers 1 to size
  for i in range(_number_):
    sheet.cell(row=1,   column=i+2).value = i+1
    sheet.cell(row=i+2, column=1  ).value = i+1

  # sheet.row_dimensions[1].font = fontObj1
  # sheet.column_dimensions['A'].font = fontObj1 There's a possible bug here
  #                                              this string causes the column
  #                                              'A' to hide

  # write multiplication values on the corresponding cells in the table
  while not it.finished:
    sheet.cell(row=it.multi_index[0]+2,column=it.multi_index[1]+2).value = \
    float(it[0])
    it.iternext()

  # Save workbook
  wb.save('multiplicationTable'+str(_number_)+'.xlsx')

myNumber = int(sys.argv[1])
createMultTable(myNumber)
