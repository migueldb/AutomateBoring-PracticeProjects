#! python3
# Automate Boring Stuff Practice Project - Chapter 12
# spreadsheet2Text.py - reads the contents of a spreadsheet and saves it, column
#                       by column to text files
#
# NOTE: This script requries a spreadsheet saved in the working directory
#

import os
import openpyxl as o

def copySS2Text(_folderPath_, _fileName_, _sheetNumber_):
    # Change directory
    os.chdir(_folderPath_)
    # Load workbook
    try:
        wb = o.load_workbook(_fileName_)
    except:
        print('Error loading the workbook')
        return     
    # Select the given worksheet number
    sheet = wb.worksheets[_sheetNumber_-1]
    # Loop through columns
    for x in range (sheet.max_column):
        # Creates a new text file for the first column in the workbook
        textFileName = str(x) + os.path.splitext(_fileName_)[0] + '.txt'
        textFile = open(textFileName, 'w')
        # Loops through rows
        for y in range(sheet.max_row):
            # If the cell is not empty write the contents to the text file
            if sheet.cell(row = y+1, column = x+1).value != None:
                textFile.write(str(sheet.cell(row = y+1, column = x+1).value) \
                               + '\n')
        # Close the text file
        textFile.close()


# User provided variables
mySheetNumber = 1
myFolderPath = os.getcwd()
myFileName = '-XTemp-Python-AutomateBoringPractice-PracticeProjects-Ch. 12.xlsx'
copySS2Text(myFolderPath, myFileName, mySheetNumber)
