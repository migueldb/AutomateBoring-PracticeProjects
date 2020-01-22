#! python3
# Automate Boring Stuff Practice Project - Chapter 14
# excel2CSV.py - Converts all the spreadsheets in a folder to CSV files
#
# NOTE: This script requries excel spreadsheets saved in the working directory
#       These spreadsheets can be downloaded here:
#       https://nostarch.com/download/
#       Automate_the_Boring_Stuff_onlinematerials_v.2.zip
#       This script requres OpenPyXl module installed.  See:
#       https://pypi.org/project/openpyxl/
#       for more information

import os, csv
import openpyxl as o

# This function converts only one excel spreadsheet in __folderPath__ with
# _fileName_ to csv
def copyOneSS2CSV(__folderPath__, _fileName_):
  # Change working directory to __folderPath__
  os.chdir(__folderPath__)
  # Create workbook object
  wb = o.load_workbook(_fileName_)
  # Loop through worksheets in workbook
  for sheetName in wb.sheetnames:
    sheet = wb[sheetName]
    # Create csv file using a the same base name but different extension
    csvFileName = os.path.splitext(_fileName_)[0] + '_' + sheetName + '.csv'
    # Create csv file object
    csvFileObj = open(csvFileName, 'w', newline='')
    # Create csv file writer object
    csvWriter = csv.writer(csvFileObj)
    # Loop through rows
    for r in range(sheet.max_row):
      rowData = []
      # Loop through columns
      for c in range(sheet.max_column):
        # append cell values to the rowData list
        rowData.append(sheet.cell(row=r+1, column=c+1,).value)
      # When finished with the row write it to the csv file
      csvWriter.writerow(rowData)
    # Close csv file
    csvFileObj.close()

# This function calls the excel to csv single file converter as many times as
# there are excel spreadsheets in _folderPath_
def CopyManySS2CSV(_folderPath_):
  for excelFile in os.listdir(_folderPath_):
    # Check if the file is an excel spreadsheet
    if excelFile.endswith('.xlsx'):
      # Call the excel to csv single file converter function
      copyOneSS2CSV(_folderPath_, excelFile)
      
# User can change the folder with the spreadsheets
myFolderPath = os.getcwd()

# Function call
CopyManySS2CSV(myFolderPath)
